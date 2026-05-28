"""Writes the bundled `btp_schedule.xlsx` — a single workbook with one sheet
per tabular artefact, formatted for planner consumption.

Sheets (in tab order):
  - summary           — run-level metadata + headline KPIs
  - kpi               — full KPI table (metric / value)
  - schedule          — lot-level schedule
  - machine_view      — schedule sorted by (machine_id, start_min)
  - building_to_curing
  - aging_violations
  - infeasibilities
  - reservation_log
  - routing_cleaned
  - audit_halt        — HALT findings only
  - audit_warn        — WARN findings only

Each sheet carries:
  - Row 1: merged title bar (navy fill, white bold text).
  - Row 2: column headers (banded blue fill, white bold).
  - Row 3+: data, with cell-level conditional fills for utilisation %,
            on_time_flag, classification, violation_type, and severity.
  - Frozen panes below row 2 so title + header stay visible while scrolling.

Use `writer_excel.read_sheet(path, sheet_name)` to read a sheet back as a
DataFrame — it handles the title-row offset (`header=1`) so callers don't
need to know about the layout.

CSV, JSON, SVG and HTML outputs (`dag.json`, `bom_graph.svg`, `gantt_*.html`,
`audit_report.md`) continue to be written alongside this workbook for
downstream tooling. The workbook is the headline artefact the planner opens.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from V1.config.enums import FindingSeverity
from V1.config.settings import Settings
from V1.models.kpi import KpiResult
from V1.models.lot import LotsResult
from V1.models.schedule import ScheduleResult
from V1.models.diagnostics import DiagnosticsResult
from V1.routes.audit import AuditResult
from V1.utilities.time_math import from_minute


WORKBOOK_NAME = "btp_schedule.xlsx"

# Layout: title in Excel row 1, header in row 2, data from row 3.
TITLE_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3
# pandas `header=` arg to use when reading these workbooks back.
PD_READ_HEADER = HEADER_ROW - 1  # 0-indexed → 1


# Excel sheet names are capped at 31 chars; pick short, stable names.
SHEET_NAMES = (
    "summary",
    "otif_by_block",
    "bottlenecks",
    "unscheduled",
    "kpi",
    "schedule",
    "machine_view",
    "building_to_curing",
    "aging_violations",
    "infeasibilities",
    "reservation_log",
    "routing_cleaned",
    "audit_halt",
    "audit_warn",
)


# Human-readable titles rendered in each sheet's row 1.
_SHEET_TITLES: dict[str, str] = {
    "summary": "Executive Summary",
    "otif_by_block": "OTIF by Curing Block (why each block is OK / LATE)",
    "bottlenecks": "Machine Bottlenecks (utilisation, ranked)",
    "unscheduled": "Unscheduled Lots + Downstream Impact",
    "kpi": "Key Performance Indicators",
    "schedule": "Production Schedule",
    "machine_view": "Machine-Level Schedule",
    "building_to_curing": "Building → Curing Handoff",
    "aging_violations": "Aging Window Violations",
    "infeasibilities": "Infeasible Lots",
    "reservation_log": "Reservation Log",
    "routing_cleaned": "Cleaned Routing",
    "audit_halt": "Audit — HALT Findings",
    "audit_warn": "Audit — Warnings",
}


# Colour palette (Excel-standard "Good/Neutral/Bad" tones + navy title bar).
_C_TITLE_BG = "1F4E78"
_C_TITLE_FG = "FFFFFF"
_C_HEADER_BG = "305496"
_C_HEADER_FG = "FFFFFF"
_C_GREEN_BG = "C6EFCE"
_C_YELLOW_BG = "FFEB9C"
_C_RED_BG = "FFC7CE"
_C_GRAY_BG = "D9D9D9"


def _fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_rgb)


# Utilisation % conditional bands (planner-supplied thresholds).
#   >= 90       → green   (high utilisation)
#   50  – <90   → yellow  (medium utilisation)
#   <  50       → red     (low utilisation)
def _utilisation_fill(pct: float) -> PatternFill | None:
    if pct >= 90.0:
        return _fill(_C_GREEN_BG)
    if pct >= 50.0:
        return _fill(_C_YELLOW_BG)
    return _fill(_C_RED_BG)


_CLASSIFICATION_FILL: dict[str, str] = {
    "OK": _C_GREEN_BG,
    "LATE": _C_RED_BG,
    "EARLY": _C_YELLOW_BG,
    "ZERO_QTY": _C_GRAY_BG,
}

_SEVERITY_FILL: dict[str, str] = {
    "HALT": _C_RED_BG,
    "WARN": _C_YELLOW_BG,
}


def read_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    """Read a sheet back as a DataFrame, accounting for the row-1 title bar.

    All downstream code (tests, planner notebooks, downstream tooling)
    should use this helper rather than `pd.read_excel` directly, so the
    title-row offset is encapsulated.
    """
    return pd.read_excel(workbook_path, sheet_name=sheet_name,
                          header=PD_READ_HEADER)


# --- builders for each sheet ----------------------------------------------

def _gt_tyre_supply(
    schedule: ScheduleResult, diag: DiagnosticsResult, settings: Settings,
) -> dict[str, float]:
    """Tyres delivered on-time vs late, derived from GT lot qty × block
    classification. One GT lot serves one curing block (L1)."""
    cls_by_block = {
        r.block_id: r.classification for r in diag.building_to_curing
    }
    on_time = late = early = zero = 0.0
    blocks_on_time = blocks_late = 0
    for s in schedule.scheduled:
        if s.item_code != settings.green_tyre_code:
            continue
        if s.qty <= 0:
            continue
        block = s.serves_blocks[0] if s.serves_blocks else None
        cls = cls_by_block.get(block, "")
        if cls == "OK":
            on_time += s.qty
            blocks_on_time += 1
        elif cls == "LATE":
            late += s.qty
            blocks_late += 1
        elif cls == "EARLY":
            early += s.qty
        else:
            zero += s.qty
    return {
        "tyres_on_time": on_time,
        "tyres_late": late,
        "tyres_early": early,
        "blocks_on_time": blocks_on_time,
        "blocks_late": blocks_late,
    }


def _busiest_machine(kpi: KpiResult) -> tuple[str, float] | None:
    """(machine_id, util_pct) of the highest-utilisation machine."""
    if not kpi.machines:
        return None
    top = max(kpi.machines, key=lambda m: m.utilisation_pct)
    return (top.machine_id, top.utilisation_pct)


def _top_consumer_of_machine(
    schedule: ScheduleResult, machine_id: str,
) -> tuple[str, float] | None:
    """(item_code, pct_of_machine_busy_min) hogging the most time on a machine."""
    by_item: dict[str, int] = {}
    total = 0
    for s in schedule.scheduled:
        if s.machine_id != machine_id or s.duration_min <= 0:
            continue
        by_item[s.item_code] = by_item.get(s.item_code, 0) + s.duration_min
        total += s.duration_min
    if not by_item or total == 0:
        return None
    item = max(by_item, key=lambda k: by_item[k])
    return (item, round(by_item[item] / total * 100, 1))


def _summary_df(
    audit: AuditResult, lots: LotsResult, schedule: ScheduleResult,
    diag: DiagnosticsResult, kpi: KpiResult, settings: Settings,
    t0: datetime, run_id: str,
) -> pd.DataFrame:
    """Sectioned executive summary — section / metric / value.

    Sections, top to bottom:
      1. RUN METADATA          — what was run
      2. LOT PRODUCTION FUNNEL — made → scheduled → not-scheduled → on-time/late
      3. GREEN TYRE SUPPLY     — curing demand fulfilment (OTIF, tyres on time)
      4. BOTTLENECK            — busiest machine + what hogs it
      5. DATA QUALITY          — audit findings + aging violations
    """
    n_sized = len(lots.lots)
    n_sched = len(schedule.scheduled)
    n_infeasible = len(schedule.infeasibilities)
    n_late_lots = sum(1 for s in schedule.scheduled if not s.on_time_flag)
    n_ontime_lots = n_sched - n_late_lots

    supply = _gt_tyre_supply(schedule, diag, settings)
    tyres_demand = settings.total_demand_tyres
    tyres_ot = supply["tyres_on_time"]

    busiest = _busiest_machine(kpi)
    span_h = round(kpi.schedule_span_min / 60.0, 1)

    rows: list[tuple[str, str, object]] = []

    def add(section: str, metric: str, value: object) -> None:
        rows.append((section, metric, value))

    # 1. RUN METADATA
    add("RUN METADATA", "Run ID", run_id)
    add("RUN METADATA", "Global start (t0)", t0.isoformat(sep=" "))
    add("RUN METADATA", "SKU", settings.sku_code)
    add("RUN METADATA", "Description", settings.sku_description)
    add("RUN METADATA", "Green tyre", settings.green_tyre_code)
    add("RUN METADATA", "Curing press", settings.curing_press)
    add("RUN METADATA", "Curing horizon",
        f"{settings.horizon_start.isoformat(sep=' ')} → "
        f"{settings.horizon_end.isoformat(sep=' ')}")
    add("RUN METADATA", "Schedule span", f"{kpi.schedule_span_min} min ({span_h} h)")

    # 2. LOT PRODUCTION FUNNEL
    add("LOT PRODUCTION FUNNEL", "Total lots made (sized)", n_sized)
    add("LOT PRODUCTION FUNNEL", "Lots scheduled (committed)",
        f"{n_sched}  ({_pct(n_sched, n_sized)}%)")
    add("LOT PRODUCTION FUNNEL", "Lots NOT scheduled (infeasible)", n_infeasible)
    add("LOT PRODUCTION FUNNEL", "  → on-time lots", n_ontime_lots)
    add("LOT PRODUCTION FUNNEL", "  → late lots (aging breach)", n_late_lots)
    add("LOT PRODUCTION FUNNEL", "Total processing minutes", kpi.total_processing_min)

    # 3. GREEN TYRE SUPPLY
    add("GREEN TYRE SUPPLY", "Curing blocks (demand events)",
        kpi.building_lots_ok + kpi.building_lots_late + kpi.building_lots_early)
    add("GREEN TYRE SUPPLY", "Blocks delivered ON TIME", kpi.building_lots_ok)
    add("GREEN TYRE SUPPLY", "Blocks delivered LATE", kpi.building_lots_late)
    add("GREEN TYRE SUPPLY", "Blocks delivered EARLY", kpi.building_lots_early)
    add("GREEN TYRE SUPPLY", "OTIF %", f"{round(kpi.otif_pct, 1)}%")
    add("GREEN TYRE SUPPLY", "Tyres demanded (total)", tyres_demand)
    add("GREEN TYRE SUPPLY", "Tyres deliverable ON TIME", int(tyres_ot))
    add("GREEN TYRE SUPPLY", "Tyres LATE", int(supply["tyres_late"]))
    add("GREEN TYRE SUPPLY", "On-time tyre %",
        f"{_pct(tyres_ot, tyres_ot + supply['tyres_late'])}%")

    # 4. BOTTLENECK
    if busiest is not None:
        mid, util = busiest
        add("BOTTLENECK", "Busiest machine", f"{mid}  ({round(util, 1)}% utilised)")
        top = _top_consumer_of_machine(schedule, mid)
        if top is not None:
            add("BOTTLENECK", f"Top consumer of {mid}",
                f"{top[0]}  ({top[1]}% of its time)")
    add("BOTTLENECK", "Changeover minutes (V1)", kpi.changeover_min)

    # 5. DATA QUALITY
    add("DATA QUALITY", "Audit HALT findings", len(audit.halt_findings))
    add("DATA QUALITY", "Audit WARN findings", len(audit.warn_findings))
    add("DATA QUALITY", "Aging-window violations", len(diag.aging_violations))
    add("DATA QUALITY", "Lot-sizing under-min warnings", len(lots.warnings))

    return pd.DataFrame(rows, columns=["section", "metric", "value"])


def _pct(num: float, den: float) -> float:
    return round(num / den * 100, 1) if den else 0.0


def _kpi_df(kpi: KpiResult) -> pd.DataFrame:
    rows: list[tuple[str, object]] = [
        ("total_lots_scheduled", kpi.total_lots_scheduled),
        ("total_lots_infeasible", kpi.total_lots_infeasible),
        ("total_lot_sizing_warnings", kpi.total_warnings),
        ("building_lots_ok", kpi.building_lots_ok),
        ("building_lots_late", kpi.building_lots_late),
        ("building_lots_early", kpi.building_lots_early),
        ("otif_pct", round(kpi.otif_pct, 3)),
        ("aging_violations_total", kpi.aging_violations_total),
        ("aging_violations_min_breaches", kpi.aging_violations_min),
        ("aging_violations_max_breaches", kpi.aging_violations_max),
        ("total_processing_min", kpi.total_processing_min),
        ("changeover_min_v1", kpi.changeover_min),
        ("schedule_span_min", kpi.schedule_span_min),
    ]
    for m in kpi.machines:
        rows.append((f"machine_{m.machine_id}_busy_min", m.busy_min))
        rows.append((f"machine_{m.machine_id}_span_min", m.span_min))
        rows.append((f"machine_{m.machine_id}_util_pct", round(m.utilisation_pct, 3)))
    return pd.DataFrame(rows, columns=["metric", "value"])


def _schedule_df(schedule: ScheduleResult, t0: datetime) -> pd.DataFrame:
    rows = []
    for s in schedule.scheduled:
        rows.append({
            "lot_id": s.lot_id,
            "item_code": s.item_code,
            "item_type": s.item_type,
            "op_seq": s.op_seq,
            "machine_id": s.machine_id,
            "start_min": s.start_min,
            "end_min": s.end_min,
            "duration_min": s.duration_min,
            "qty": s.qty,
            "uom": s.uom,
            "serves_blocks": "|".join(s.serves_blocks),
            "on_time_flag": s.on_time_flag,
            "start_dt": from_minute(s.start_min, t0),
            "end_dt": from_minute(s.end_min, t0),
        })
    cols = ["lot_id", "item_code", "item_type", "op_seq", "machine_id",
            "start_min", "end_min", "duration_min", "qty", "uom",
            "serves_blocks", "on_time_flag", "start_dt", "end_dt"]
    return pd.DataFrame(rows, columns=cols)


def _machine_view_df(schedule_df: pd.DataFrame) -> pd.DataFrame:
    if schedule_df.empty:
        return schedule_df.copy()
    return schedule_df.sort_values(
        ["machine_id", "start_min", "lot_id"], kind="stable"
    ).reset_index(drop=True)


def _btc_df(diag: DiagnosticsResult) -> pd.DataFrame:
    rows = [
        {
            "lot_id": r.lot_id, "machine_id": r.machine_id,
            "block_id": r.block_id,
            "gt_end_min": r.gt_end_min,
            "curing_start_min": r.curing_start_min,
            "gap_min": r.gap_min,
            "min_aging_min": r.min_aging_min,
            "max_aging_min": r.max_aging_min,
            "classification": r.classification,
        }
        for r in diag.building_to_curing
    ]
    cols = ["lot_id", "machine_id", "block_id", "gt_end_min",
            "curing_start_min", "gap_min", "min_aging_min", "max_aging_min",
            "classification"]
    return pd.DataFrame(rows, columns=cols)


def _aging_violations_df(diag: DiagnosticsResult) -> pd.DataFrame:
    rows = [
        {
            "consumer_lot": v.consumer_lot_id,
            "predecessor_lot": v.producer_lot_id,
            "item_code": v.item_code,
            "edge_min": v.edge_min,
            "edge_max": v.edge_max,
            "actual_gap": v.actual_gap_min,
            "violation_type": v.violation_type,
        }
        for v in diag.aging_violations
    ]
    cols = ["consumer_lot", "predecessor_lot", "item_code", "edge_min",
            "edge_max", "actual_gap", "violation_type"]
    return pd.DataFrame(rows, columns=cols)


def _infeasibilities_df(schedule: ScheduleResult) -> pd.DataFrame:
    rows = [
        {
            "lot_id": i.lot_id, "item_code": i.item_code,
            "op_seq": i.op_seq,
            "binding_constraint": i.binding_constraint,
            "message": i.message,
        }
        for i in schedule.infeasibilities
    ]
    cols = ["lot_id", "item_code", "op_seq", "binding_constraint", "message"]
    return pd.DataFrame(rows, columns=cols)


def _otif_by_block_df(
    diag: DiagnosticsResult, schedule: ScheduleResult, settings: Settings,
) -> pd.DataFrame:
    """Per-curing-block OTIF detail — the WHY behind each LATE block.

    For each Building→Curing record we add the tyre qty and, for LATE
    blocks, the binding component (the GT lot's latest-finishing producer).
    """
    # Index GT lots by block + the latest producer per GT lot.
    gt_by_block: dict[str, ScheduledLot] = {}
    sched_by_id = {s.lot_id: s for s in schedule.scheduled}
    for s in schedule.scheduled:
        if s.item_code == settings.green_tyre_code and s.serves_blocks:
            gt_by_block[s.serves_blocks[0]] = s

    def _binding_component(gt: ScheduledLot) -> str:
        """Item whose producer finished latest — the reason GT couldn't
        start sooner."""
        latest_end = -1
        latest_item = ""
        for ing, prod_ids in gt.producer_lot_ids.items():
            for pid in prod_ids:
                p = sched_by_id.get(pid)
                if p is not None and p.end_min > latest_end:
                    latest_end = p.end_min
                    latest_item = ing
        return latest_item

    rows = []
    for r in diag.building_to_curing:
        gt = gt_by_block.get(r.block_id)
        tyres = int(gt.qty) if gt is not None else 0
        binding = ""
        if r.classification == "LATE" and gt is not None:
            binding = _binding_component(gt)
        rows.append({
            "block_id": r.block_id,
            "tyres": tyres,
            "curing_start_min": r.curing_start_min,
            "gt_end_min": r.gt_end_min,
            "gap_min": r.gap_min,
            "classification": r.classification,
            "binding_component": binding,
        })
    cols = ["block_id", "tyres", "curing_start_min", "gt_end_min",
            "gap_min", "classification", "binding_component"]
    return pd.DataFrame(rows, columns=cols)


def _bottlenecks_df(
    kpi: KpiResult, schedule: ScheduleResult,
) -> pd.DataFrame:
    """Machine utilisation ranked high→low, with the item hogging each
    machine. The top rows are the binding resources for the schedule span."""
    rows = []
    for m in sorted(kpi.machines, key=lambda x: -x.utilisation_pct):
        top = _top_consumer_of_machine(schedule, m.machine_id)
        lots_on = sum(1 for s in schedule.scheduled
                      if s.machine_id == m.machine_id and s.duration_min > 0)
        rows.append({
            "machine_id": m.machine_id,
            "lots": lots_on,
            "busy_min": m.busy_min,
            "span_min": m.span_min,
            "utilisation_pct": round(m.utilisation_pct, 1),
            "top_item": top[0] if top else "",
            "top_item_pct_of_machine": top[1] if top else 0.0,
        })
    cols = ["machine_id", "lots", "busy_min", "span_min", "utilisation_pct",
            "top_item", "top_item_pct_of_machine"]
    return pd.DataFrame(rows, columns=cols)


def _unscheduled_df(
    schedule: ScheduleResult, lots: LotsResult, settings: Settings,
) -> pd.DataFrame:
    """Not-scheduled (infeasible) lots + the downstream curing blocks each
    one would have fed. Empty when every lot committed (the common case
    under L11 flag-and-continue)."""
    serves_by_lot = {l.lot_id: l.serves_blocks for l in lots.lots}
    rows = []
    for i in schedule.infeasibilities:
        blocks = serves_by_lot.get(i.lot_id, [])
        rows.append({
            "lot_id": i.lot_id,
            "item_code": i.item_code,
            "op_seq": i.op_seq,
            "binding_constraint": i.binding_constraint,
            "downstream_blocks_affected": "|".join(blocks) if blocks else "—",
            "n_blocks_affected": len(blocks),
            "message": i.message,
        })
    cols = ["lot_id", "item_code", "op_seq", "binding_constraint",
            "downstream_blocks_affected", "n_blocks_affected", "message"]
    return pd.DataFrame(rows, columns=cols)


def _reservation_log_df(schedule: ScheduleResult) -> pd.DataFrame:
    rows = [
        {
            "event_minute": e.event_minute,
            "event_type": e.event_type,
            "consumer_lot_id": e.consumer_lot_id,
            "producer_lot_id": e.producer_lot_id,
            "item_code": e.item_code,
            "qty": e.qty,
            "producer_end_min": e.producer_end_min,
            "latest_acceptable_start_min": e.latest_acceptable_start_min,
        }
        for e in schedule.reservation_log
    ]
    cols = ["event_minute", "event_type", "consumer_lot_id",
            "producer_lot_id", "item_code", "qty", "producer_end_min",
            "latest_acceptable_start_min"]
    return pd.DataFrame(rows, columns=cols)


def _routing_cleaned_df(audit: AuditResult) -> pd.DataFrame:
    # Drop the in-memory list column for CSV/Excel friendliness.
    return audit.routing_cleaned_df.drop(columns=["machines_list"], errors="ignore")


def _findings_df(audit: AuditResult, severity: FindingSeverity) -> pd.DataFrame:
    findings = [f for f in audit.findings if f.severity == severity]
    rows = []
    for f in findings:
        rows.append({
            "severity": f.severity.value,
            "code": f.code,
            "sheet": f.sheet or "",
            "source_row_pandas": f.source_row if f.source_row is not None else "",
            "source_row_excel": f.excel_row() if f.excel_row() is not None else "",
            "item_code": f.item_code or "",
            "message": f.message,
        })
    cols = ["severity", "code", "sheet", "source_row_pandas",
            "source_row_excel", "item_code", "message"]
    return pd.DataFrame(rows, columns=cols)


# --- public entry ---------------------------------------------------------

def write_full(
    audit: AuditResult,
    lots: LotsResult,
    schedule: ScheduleResult,
    diag: DiagnosticsResult,
    kpi: KpiResult,
    settings: Settings,
    t0: datetime,
    run_id: str,
    output_dir: Path,
) -> Path:
    """Build and write the full bundled workbook. Returns its path."""
    output_dir.mkdir(parents=True, exist_ok=True)
    sched_df = _schedule_df(schedule, t0)
    sheets: dict[str, pd.DataFrame] = {
        "summary": _summary_df(audit, lots, schedule, diag, kpi,
                               settings, t0, run_id),
        "otif_by_block": _otif_by_block_df(diag, schedule, settings),
        "bottlenecks": _bottlenecks_df(kpi, schedule),
        "unscheduled": _unscheduled_df(schedule, lots, settings),
        "kpi": _kpi_df(kpi),
        "schedule": sched_df,
        "machine_view": _machine_view_df(sched_df),
        "building_to_curing": _btc_df(diag),
        "aging_violations": _aging_violations_df(diag),
        "infeasibilities": _infeasibilities_df(schedule),
        "reservation_log": _reservation_log_df(schedule),
        "routing_cleaned": _routing_cleaned_df(audit),
        "audit_halt": _findings_df(audit, FindingSeverity.HALT),
        "audit_warn": _findings_df(audit, FindingSeverity.WARN),
    }
    return _write_workbook(sheets, output_dir)


def write_halt(
    audit: AuditResult,
    settings: Settings,
    t0: datetime,
    run_id: str,
    output_dir: Path,
) -> Path:
    """Write a HALT-only workbook (audit ran, downstream did not).

    Contains just the summary, routing_cleaned, and the two audit-findings
    sheets so the planner can diagnose without leaving Excel.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_rows = [
        ("run_id", run_id),
        ("t0", t0.isoformat(sep=" ")),
        ("sku_code", settings.sku_code),
        ("status", "HALT"),
        ("audit_halt_findings", len(audit.halt_findings)),
        ("audit_warn_findings", len(audit.warn_findings)),
        ("schedule_emitted", "NO — audit HALT blocked downstream"),
    ]
    sheets: dict[str, pd.DataFrame] = {
        "summary": pd.DataFrame(summary_rows, columns=["metric", "value"]),
        "audit_halt": _findings_df(audit, FindingSeverity.HALT),
        "audit_warn": _findings_df(audit, FindingSeverity.WARN),
        "routing_cleaned": _routing_cleaned_df(audit),
    }
    return _write_workbook(sheets, output_dir)


def _write_workbook(sheets: dict[str, pd.DataFrame], output_dir: Path) -> Path:
    path = output_dir / WORKBOOK_NAME
    # `startrow=HEADER_ROW - 1` lays the pandas-emitted header on Excel
    # row 2; row 1 is left empty for the merged title bar.
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False,
                         startrow=HEADER_ROW - 1)
            ws = writer.sheets[name]
            _format_sheet(ws, df, name)
    return path


def _format_sheet(ws, df: pd.DataFrame, sheet_name: str) -> None:
    """Apply title bar, header banding, frozen panes, column widths, and
    per-sheet categorical fills. Pure presentation — no data mutation."""
    n_cols = max(1, len(df.columns))
    n_rows = len(df)

    # --- Row 1: merged title bar ------------------------------------------
    title = _SHEET_TITLES.get(
        sheet_name, sheet_name.replace("_", " ").title()
    )
    ws.cell(row=TITLE_ROW, column=1, value=title)
    if n_cols > 1:
        ws.merge_cells(
            start_row=TITLE_ROW, end_row=TITLE_ROW,
            start_column=1, end_column=n_cols,
        )
    title_cell = ws.cell(row=TITLE_ROW, column=1)
    title_cell.font = Font(size=14, bold=True, color=_C_TITLE_FG)
    title_cell.fill = _fill(_C_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 28

    # --- Row 2: column headers --------------------------------------------
    header_font = Font(bold=True, color=_C_HEADER_FG)
    header_fill = _fill(_C_HEADER_BG)
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[HEADER_ROW].height = 22

    # Freeze title + header so they stay visible while scrolling data.
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=1)

    # --- Per-sheet categorical fills (data rows) --------------------------
    if sheet_name == "kpi" and n_rows > 0:
        _shade_kpi_utilisation(ws, df)
    elif sheet_name in ("schedule", "machine_view") and n_rows > 0:
        _shade_on_time_flag(ws, df)
    elif sheet_name in ("building_to_curing", "otif_by_block") and n_rows > 0:
        _shade_classification(ws, df)
    elif sheet_name == "aging_violations" and n_rows > 0:
        _shade_violation_type(ws, df)
    elif sheet_name in ("audit_halt", "audit_warn") and n_rows > 0:
        _shade_severity(ws, df)
    elif sheet_name == "summary" and n_rows > 0:
        _shade_summary_sections(ws, df)
    elif sheet_name == "bottlenecks" and n_rows > 0:
        _shade_bottleneck_util(ws, df)

    # --- Column widths (header text length + 2, capped at 60) -------------
    for col_idx, col in enumerate(df.columns, start=1):
        width = min(60, max(12, len(str(col)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _data_range(n_rows: int) -> range:
    """Excel row indices (1-based) for the data block."""
    return range(DATA_START_ROW, DATA_START_ROW + n_rows)


def _col_index(df: pd.DataFrame, col_name: str) -> int | None:
    """1-based column index of `col_name` in df, or None if absent."""
    if col_name not in df.columns:
        return None
    return list(df.columns).index(col_name) + 1


def _shade_kpi_utilisation(ws, df: pd.DataFrame) -> None:
    """Colour every utilisation-% cell in the kpi sheet.

    The kpi sheet is a (metric, value) long-table; we scan for rows whose
    metric name ends in `_util_pct` and shade the value cell.
    """
    metric_col = _col_index(df, "metric")
    value_col = _col_index(df, "value")
    if metric_col is None or value_col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        metric = str(df.at[df_row_idx, "metric"])
        if not metric.endswith("_util_pct"):
            continue
        raw = df.at[df_row_idx, "value"]
        try:
            pct = float(raw)
        except (TypeError, ValueError):
            continue
        fill = _utilisation_fill(pct)
        if fill is not None:
            ws.cell(row=row_idx, column=value_col).fill = fill


def _shade_on_time_flag(ws, df: pd.DataFrame) -> None:
    col = _col_index(df, "on_time_flag")
    if col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        val = df.at[df_row_idx, "on_time_flag"]
        if val is True or str(val).strip().lower() == "true":
            ws.cell(row=row_idx, column=col).fill = _fill(_C_GREEN_BG)
        elif val is False or str(val).strip().lower() == "false":
            ws.cell(row=row_idx, column=col).fill = _fill(_C_RED_BG)


def _shade_classification(ws, df: pd.DataFrame) -> None:
    col = _col_index(df, "classification")
    if col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        cls = str(df.at[df_row_idx, "classification"])
        rgb = _CLASSIFICATION_FILL.get(cls)
        if rgb is not None:
            ws.cell(row=row_idx, column=col).fill = _fill(rgb)


# Light section-band fills for the summary sheet (rotating, easy to scan).
_SECTION_BANDS = ["EAF1FB", "E8F6EF", "FBF2E0", "F3EEFA", "FDEDED"]


def _shade_summary_sections(ws, df: pd.DataFrame) -> None:
    """Tint the `section` cell of the summary so each block is visually
    grouped. Also bold the first row of each section's section cell."""
    sec_col = _col_index(df, "section")
    if sec_col is None:
        return
    seen: dict[str, str] = {}
    band_i = 0
    prev = None
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        sec = str(df.at[df_row_idx, "section"])
        if sec not in seen:
            seen[sec] = _SECTION_BANDS[band_i % len(_SECTION_BANDS)]
            band_i += 1
        ws.cell(row=row_idx, column=sec_col).fill = _fill(seen[sec])
        # Blank the repeated section label so only the first row shows it.
        if sec == prev:
            ws.cell(row=row_idx, column=sec_col).value = ""
        else:
            ws.cell(row=row_idx, column=sec_col).font = Font(bold=True)
        prev = sec


def _shade_bottleneck_util(ws, df: pd.DataFrame) -> None:
    col = _col_index(df, "utilisation_pct")
    if col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        try:
            pct = float(df.at[df_row_idx, "utilisation_pct"])
        except (TypeError, ValueError):
            continue
        fill = _utilisation_fill(pct)
        if fill is not None:
            ws.cell(row=row_idx, column=col).fill = fill


def _shade_violation_type(ws, df: pd.DataFrame) -> None:
    """Any aging-window violation is bad — colour the type cell red."""
    col = _col_index(df, "violation_type")
    if col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        val = df.at[df_row_idx, "violation_type"]
        if isinstance(val, str) and val:
            ws.cell(row=row_idx, column=col).fill = _fill(_C_RED_BG)


def _shade_severity(ws, df: pd.DataFrame) -> None:
    col = _col_index(df, "severity")
    if col is None:
        return
    for row_idx, df_row_idx in zip(_data_range(len(df)), df.index):
        sev = str(df.at[df_row_idx, "severity"])
        rgb = _SEVERITY_FILL.get(sev)
        if rgb is not None:
            ws.cell(row=row_idx, column=col).fill = _fill(rgb)
